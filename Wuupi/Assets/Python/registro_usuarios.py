# registro_usuarios.py
from openpyxl import Workbook, load_workbook
import os

def guardar_usuario(username, password, archivo='usuarios.xlsx'):
    if os.path.exists(archivo):
        wb = load_workbook(archivo)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Usuario", "Contraseña"])

    ws.append([username, password])
    wb.save(archivo)
    print(f"Usuario {username} registrado.")